# coding=utf-8
from openpyxl import load_workbook


class Excel:

    def __init__(self, dst):
        self.dst = dst
        self.wb = load_workbook(dst)
        self.index = 1

    def _add_sheet(self, sheet_name):
        return self.wb.create_sheet(sheet_name[:31])

    def _add_method_summary(self, ws, req_class, http_method, method_title, method_path):
        ws.append(['接口类名', '接口描述', '接口路径', '请求方式', '方法描述', '其他'])
        ws.append([req_class, '', method_path, http_method, method_title, ''])

    def _req_body_loop(self, ws, req_class, req_json: dict):
        later_deal = []
        for key_name, items in req_json.items():
            if isinstance(items, tuple):
                # 如果元组数量为两个，则说明是Object
                if len(items) == 2:
                    key_class, son_req_json = items
                    ws.append([req_class, key_name, key_class, "否", "", ""])
                    later_deal.append((key_class, son_req_json))
                else:
                    # 如果不是两个，则是具体数值
                    items = items if len(items) == 6 else list(items) + [""]
                    ws.append(items)
            else:
                raise RuntimeError(items)
        for key_class, son_req_json in later_deal:
            self._req_body_loop(ws, key_class, son_req_json)

    def _add_body(self, ws, body):
        top_class, actual_body = body
        self._req_body_loop(ws, top_class, actual_body)

    def _add_request_body(self, ws, req_body):
        ws.append(['请求报文', '', '', '', '', ''])
        ws.append(['请求类', '参数编码', '返回类型', '必输', '说明', ''])
        if req_body:
            self._add_body(ws, req_body)

    def _add_response_body(self, ws, res_body):
        ws.append(['响应报文', '', '', '', '', ''])
        ws.append(['响应类', '参数编码', '返回类型', '必输', '说明', ''])
        if res_body:
            self._add_body(ws, res_body)

    def add_req_item(self, interface_name, method_name, http_method, method_title, req_body, res_json, method_path):
        ws = self._add_sheet(str(self.index) + "." + method_name)
        self._add_method_summary(ws, interface_name, http_method, method_title, method_path)
        self._add_request_body(ws, req_body)
        self._add_response_body(ws, res_json)
        self.index += 1

    def save(self):
        self.wb.save(self.dst)
        self.wb.close()
